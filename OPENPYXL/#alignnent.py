#alignnent.py

from openpyxl import Workbook
from openpyxl.styles import Alignment

def center_text(path, horizontal='center', vertical='center'):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Hello'
    ws['A1'].alignment = Alignment(horizontal=horizontal, vertical=vertical)
    wb.save(path)
    
    
    ws['A2'] = 'from'
    ws['A3'] = 'Openpyxl'
    ws['A3'].alignment = Alignment(text_rotation=90)
    wb.save(path)
    
    
    
if __name__ == '__main__':
    center_text('alignment.xlsx')
    