#editing_demo.py

from openpyxl import load_workbook


def edit(path, data):
    wb = load_workbook(filename=path)
    ws = wb.active
    for cell in data:
        current_value = ws[cell].value
        ws[cell] = data[cell]
        print(f'Changing {cell}) from {current_value} to {data[cell]}')
    wb.save(path)


if __name__ == '__main__':
    data = {
        'B1':"Hi", "B5": "Python"}
    edit("/home/q/github project repos/Github Repos/PortfolioProjects/OPENPYXL/inserting.xlsx", data)