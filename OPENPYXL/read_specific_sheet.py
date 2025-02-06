from openpyxl import load_workbook

def open_workbook(path, sheet_name):
    workbook = load_workbook(filename=path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"The title of the Worksheet is: {sheet.title}")
        print(f"Cells that contain data: {sheet.calculate_dimension()}")
    else:
        print(f"Sheet '{sheet_name}' not found in the workbook.")

if __name__ == "__main__":
    open_workbook("/home/q/github project repos/Github Repos/PortfolioProjects/OPENPYXL/Book1.xlsx", sheet_name="Budget")
