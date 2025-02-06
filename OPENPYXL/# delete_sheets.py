# delete_sheets.py 

import openpyxl


def create_worksheets(path):
    wb = openpyxl.Workbook()
    wb.create_sheet()
    print(wb.sheetnames)
    #insert a worksheet
    wb.create_sheet(index=1, title="Second Sheet")
    print(wb.sheetnames)
    del wb["Second Sheet"]
    print(wb.sheetnames)
    wb.save(path)


if __name__ == "__main__":
    create_worksheets("delete_sheets.xlsx")