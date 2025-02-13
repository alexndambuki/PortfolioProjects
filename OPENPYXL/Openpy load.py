#open_workbook.py

from openpyxl import load_workbook

def open_workbook(file_path):
    workbook = load_workbook(filename=file_path)
    print(f"Worksheet names: {workbook.sheetnames}")
    sheet = workbook.active
    print(sheet)
    print(f"The title of the Worksheet is: {sheet.title}")

if __name__ == "__main__":
    open_workbook("C:\\PROJECTS\\Github Repos\\PortfolioProjects\\OPENPYXL\\Book1.xlsx")
