#reading_column_cells.py

from openpyxl import load_workbook


def iterating_column(path, sheet_name, col):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return

    sheet = workbook[sheet_name]
    for cell in sheet[col]:
        print(f"{cell.column_letter}{cell.row} = {cell.value}")
 

if __name__ == "__main__":
    iterating_column("/home/q/github project repos/Github Repos/PortfolioProjects/OPENPYXL/Book1.xlsx", sheet_name="Activities", 
                     col="A")