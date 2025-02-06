# iterating_over_cells_in_rows.py

from openpyxl import load_workbook
import openpyxl

def iterating_over_values(path, sheet_name, cell_range):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found Quitting...")
        return
    
    sheet = workbook[sheet_name]
    for column in sheet[cell_range]:
        for cell in sheet[cell_range]:
            for cell in column:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    #Skip this cell
                    continue
                print(f"{cell.column_letter}{cell.row} = {cell.value}")


if __name__ == "__main__":
    iterating_over_values("Book1.xlsx", "Activities", cell_range="A1:C5")