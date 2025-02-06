#creating_spreadsheet.py
from openpyxl import Workbook

def create_Workbook(path):
    workbook = Workbook()
    workbook.save(path)

if __name__ == '__main__':
    create_Workbook('hello.xlsx')