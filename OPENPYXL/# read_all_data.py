# read_all_data.py

#import openpyxl
#from openpyxl import load_workbook


#def read_all_data(path):
#    workbook = load_workbook(filename=path)
#    for sheet_name in workbook.sheetnames:
#        sheet = workbook[sheet_name]
#        print(f"Title = {sheet.title}")
#        for row in sheet.rows:
#            for cell in row:
#                if isinstance(cell, openpyxl.cell.cell.MergedCell):
#                    #skip this cell
#                    continue
#                print(f"{cell.column_letter}{cell.row} = {cell.value}") 



#if __name__ == "__main__":
#    read_all_data("/home/q/github project repos/Github Repos/PortfolioProjects/OPENPYXL/Book1.xlsx")

import openpyxl
from openpyxl import load_workbook


def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Title = {sheet.title}")
        for value in sheet.iter_rows(values_only=True):
            print(value)



if __name__ == "__main__":
    read_all_data("/home/q/github project repos/Github Repos/PortfolioProjects/OPENPYXL/Book1.xlsx")
